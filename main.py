from openpyxl import load_workbook

# file_path = 'pyscp.xlsx' # your exl file here

# wb = load_workbook(file_path)
# ws = wb.active

# USE THESE VALUES FOR TESTING
# self.ws.insert_cols(ending_col)
# self.ws['A1'] = "harvey"
# self.ws['B1'] = "jiang"
# self.ws['A2'] = "wilson"
# self.ws['B2'] = "boooochen"
# self.ws['A3'] = "rayn"
# self.ws['B3'] = "zanhe"


class Excel:
    def __init__(self, file_path_excel: str, final_name_row: int = 1):
        self.wb = load_workbook(file_path_excel)
        self.ws = self.wb.active
        self.final_name_row = final_name_row
        self.file_path_excel = file_path_excel

    def concat_names(self, first_initial_row: int = 0, last_name_row: int = 1, max_rows: int= 4, ending_col: int = 3):

        for index, row in enumerate(self.ws.iter_rows(values_only=True)):
           
            first_initial = row[first_initial_row]
            last_name = row[last_name_row]
            combined_name = [first_initial,'.', last_name]
            end_string = ''.join(combined_name)
            self.ws[f"C{self.final_name_row}"] = end_string
            self.final_name_row += 1

        self.ws.delete_cols(max_rows)
        self.wb.save(self.file_path_excel)

    def test_concat(self, min_row: int = 1, max_row: int = 3, values_only = True):

        for row in self.ws.iter_rows(min_row=min_row, max_row=max_row, values_only=values_only):  # test
            print(row)


test = Excel('pyscp.xlsx')
test.concat_names()
# print(test.test_concat())


